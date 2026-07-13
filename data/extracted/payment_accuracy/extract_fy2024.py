#!/usr/bin/env python3
"""Extract exact FY2024 PaymentAccuracy tables with source boundaries intact."""

import json
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[3]
RAW = ROOT / "data/raw/omb/SRC-OMB-PAYMENTACCURACY/2026-07-12/FY2024_Dataset.xlsx"
OUT = Path(__file__).resolve().parent
SOURCE_ID = "SRC-OMB-PAYMENTACCURACY-FY2024-DATA"


def clean(value):
    if isinstance(value, float):
        return round(value, 9)
    return value


def dict_rows(ws, header_row, start_row):
    headers = [str(value).strip() if value is not None else None for value in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]
    for row_number, values in enumerate(ws.iter_rows(min_row=start_row, values_only=True), start=start_row):
        if not any(value is not None for value in values):
            continue
        yield row_number, {header:clean(value) for header, value in zip(headers, values) if header is not None}


def write_jsonl(path, rows):
    path.write_text("".join(json.dumps(row, separators=(",", ":")) + "\n" for row in rows), encoding="utf-8")


def main():
    workbook = load_workbook(RAW, read_only=True, data_only=True)
    program_rows = []
    for source_row, row in dict_rows(workbook["All Program Results"], 1, 2):
        if str(row.get("Fiscal Year")) != "2024":
            continue
        row.update({"record_family":"payment_accuracy_program_result", "source_id":SOURCE_ID, "source_sheet":"All Program Results", "source_row":source_row, "extraction_status":"draft_extracted"})
        program_rows.append(row)
    write_jsonl(OUT / "fy2024_program_results.v1.draft.jsonl", program_rows)

    total_row = None
    for source_row, row in dict_rows(workbook["Improper Payment Totals"], 1, 2):
        if str(row.get("Fiscal Year")) == "2024":
            total_row = row
            total_row.update({"record_family":"payment_accuracy_governmentwide_total", "source_id":SOURCE_ID, "source_sheet":"Improper Payment Totals", "source_row":source_row, "extraction_status":"draft_extracted"})
            break
    if total_row is None:
        raise ValueError("FY2024 government-wide total missing")
    (OUT / "fy2024_governmentwide_total.v1.draft.json").write_text(json.dumps(total_row, indent=2) + "\n", encoding="utf-8")

    fraud_rows = []
    for source_row, row in dict_rows(workbook["Confirmed Fraud"], 4, 5):
        if str(row.get("Fiscal Year")) != "2024":
            continue
        row.update({"record_family":"payment_accuracy_confirmed_fraud", "source_id":SOURCE_ID, "source_sheet":"Confirmed Fraud", "source_row":source_row, "definition":"court-confirmed cases only; excludes out-of-court settlements with or without admission of guilt", "extraction_status":"draft_extracted"})
        fraud_rows.append(row)
    write_jsonl(OUT / "fy2024_confirmed_fraud.v1.draft.jsonl", fraud_rows)

    recovery_rows = []
    for source_row, row in dict_rows(workbook["Recovery Details "], 4, 5):
        if str(row.get("Fiscal Year")) != "2024":
            continue
        row.update({"record_family":"payment_accuracy_agency_recovery", "source_id":SOURCE_ID, "source_sheet":"Recovery Details", "source_row":source_row, "scope_note":"agency-level recovery activity; not a direct subset of estimated program overpayments", "extraction_status":"draft_extracted"})
        recovery_rows.append(row)
    write_jsonl(OUT / "fy2024_agency_recovery.v1.draft.jsonl", recovery_rows)

    over = total_row["Total Overpayment Amount ($M)"]
    under = total_row["Underpayment Amount ($M)"]
    technical = total_row["Technically Improper Payment Amount ($M)"]
    improper = total_row["Improper Payment Amount ($M)"]
    unknown = total_row["Unknown Payment Amount ($M)"]
    combined = total_row["Improper Payment and Unknown Payment Amount ($M)"]
    if abs((over + under + technical) - improper) > 0.001:
        raise ValueError("improper payment classes do not reconcile")
    if abs((improper + unknown) - combined) > 0.001:
        raise ValueError("improper plus unknown does not reconcile")
    summary = {
        "record_id":"payment-accuracy-fy2024-extraction-summary",
        "source_id":SOURCE_ID,
        "program_result_rows":len(program_rows),
        "confirmed_fraud_rows":len(fraud_rows),
        "agency_recovery_rows":len(recovery_rows),
        "reconciliation":{"overpayment_plus_underpayment_plus_technical_equals_improper":True,"improper_plus_unknown_equals_combined":True},
        "status":"draft_extracted_not_public_claim"
    }
    (OUT / "fy2024_extraction_summary.v1.draft.json").write_text(json.dumps(summary, indent=2) + "\n", encoding="utf-8")
    print(f"extracted {len(program_rows)} program rows, {len(fraud_rows)} confirmed-fraud rows, and {len(recovery_rows)} recovery rows")


if __name__ == "__main__":
    main()
